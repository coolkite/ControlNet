import os
import re
import statistics
import openpyxl
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import numpy as np

# Hyperparameters
INTERVALS = [0,18,19,20,21,22,23,24,25,50]
VIEWS = [0, 4, 8, 12, 16]
METRICS = ['MSE', 'IoU']
MASK_TYPES = ['a_original_depth_mask', 'c_baseline_dino_mask', 'd_learned_dino_mask']

def extract_metric(file_path, metric):
    with open(file_path, 'r') as f:
        content = f.read()
        match = re.search(f'{metric}: ([\d.]+)', content)
        return float(match.group(1)) if match else None

def process_data(base_dir):
    data = {interval: {view: {metric.lower(): {} for metric in METRICS} 
            for view in VIEWS} 
            for interval in INTERVALS}
    
    for root, _, files in os.walk(base_dir):
        if 'baseline_metrics.txt' in files or 'learned_metrics.txt' in files:
            parts = root.split(os.sep)
            mask_type = parts[-6]
            interval = int(parts[-3])
            view = int(parts[-2])
            chair_id = parts[-4]
            
            if interval not in INTERVALS or view not in VIEWS:
                continue
            
            if mask_type == MASK_TYPES[0]:  # a_original_depth_mask
                baseline_file = os.path.join(root, 'baseline_metrics.txt')
                learned_file = os.path.join(root, 'learned_metrics.txt')
                
                baseline_mse = extract_metric(baseline_file, 'MSE')
                learned_mse = extract_metric(learned_file, 'MSE')
                
                if chair_id not in data[interval][view]['mse']:
                    data[interval][view]['mse'][chair_id] = {'baseline': [], 'learned': []}
                
                data[interval][view]['mse'][chair_id]['baseline'].append(baseline_mse)
                data[interval][view]['mse'][chair_id]['learned'].append(learned_mse)
            
            elif mask_type == MASK_TYPES[1]:  # c_baseline_dino_mask
                baseline_iou_file = os.path.join(root, 'baseline_metrics.txt')
                baseline_iou = extract_metric(baseline_iou_file, 'IoU')
                
                if chair_id not in data[interval][view]['iou']:
                    data[interval][view]['iou'][chair_id] = {'baseline': [], 'learned': []}
                
                data[interval][view]['iou'][chair_id]['baseline'].append(baseline_iou)
            
            elif mask_type == MASK_TYPES[2]:  # d_learned_dino_mask
                learned_iou_file = os.path.join(root, 'learned_metrics.txt')
                learned_iou = extract_metric(learned_iou_file, 'IoU')
                
                if chair_id not in data[interval][view]['iou']:
                    data[interval][view]['iou'][chair_id] = {'baseline': [], 'learned': []}
                
                data[interval][view]['iou'][chair_id]['learned'].append(learned_iou)
    
    return data

def calculate_stats(data):
    stats = {interval: {view: {metric.lower(): {} for metric in METRICS} 
             for view in VIEWS} 
             for interval in INTERVALS}
    
    for interval in data:
        for view in data[interval]:
            for metric in METRICS:
                for chair_id in data[interval][view][metric.lower()]:
                    stats[interval][view][metric.lower()][chair_id] = {}
                    for token in ['baseline', 'learned']:
                        values = data[interval][view][metric.lower()][chair_id][token]
                        if values:
                            mean = np.mean(values)
                            std = np.std(values) if len(values) > 1 else 0
                            stats[interval][view][metric.lower()][chair_id][token] = f"{mean:.2f}±{std:.2f}"
                        else:
                            stats[interval][view][metric.lower()][chair_id][token] = "N/A"
    
    return stats

def create_excel(stats, output_file):
    wb = openpyxl.Workbook()
    
    for interval in INTERVALS:
        for metric in METRICS:
            ws = wb.create_sheet(f"{metric}_{interval}")
            
            headers = [''] + sum([[f'View {view}', ''] for view in VIEWS], []) + ['All', '']
            ws.append(headers)
            
            subheaders = [''] + ['baseline', 'learned'] * (len(VIEWS) + 1)
            ws.append(subheaders)
            
            chair_ids = list(set(chair_id for view in VIEWS for chair_id in stats[interval][view][metric.lower()]))
            
            for chair_id in chair_ids:
                row = [chair_id]
                
                all_values = {'baseline': [], 'learned': []}
                for view in VIEWS:
                    if chair_id in stats[interval][view][metric.lower()]:
                        row.extend([
                            stats[interval][view][metric.lower()][chair_id]['baseline'],
                            stats[interval][view][metric.lower()][chair_id]['learned']
                        ])
                        for token in ['baseline', 'learned']:
                            if '±' in stats[interval][view][metric.lower()][chair_id][token]:
                                mean, _ = map(float, stats[interval][view][metric.lower()][chair_id][token].split('±'))
                                all_values[token].append(mean)
                    else:
                        row.extend(['N/A', 'N/A'])
                
                # Calculate 'All' column
                for token in ['baseline', 'learned']:
                    if all_values[token]:
                        mean = np.mean(all_values[token])
                        std = np.std(all_values[token]) if len(all_values[token]) > 1 else 0
                        row.append(f"{mean:.2f}±{std:.2f}")
                    else:
                        row.append('N/A')
                
                ws.append(row)
            
            # Calculate overall mean
            overall_mean_row = ['Overall mean']
            for view in VIEWS + ['All']:
                for token in ['baseline', 'learned']:
                    all_values = []
                    for chair_id in chair_ids:
                        if view == 'All':
                            for v in VIEWS:
                                if chair_id in stats[interval][v][metric.lower()]:
                                    value = stats[interval][v][metric.lower()][chair_id][token]
                                    if '±' in value:
                                        mean, _ = map(float, value.split('±'))
                                        all_values.append(mean)
                        elif chair_id in stats[interval][view][metric.lower()]:
                            value = stats[interval][view][metric.lower()][chair_id][token]
                            if '±' in value:
                                mean, _ = map(float, value.split('±'))
                                all_values.append(mean)
                    if all_values:
                        mean = np.mean(all_values)
                        std = np.std(all_values) if len(all_values) > 1 else 0
                        overall_mean_row.append(f"{mean:.2f}±{std:.2f}")
                    else:
                        overall_mean_row.append('N/A')
            ws.append(overall_mean_row)

            # Apply alignment
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    wb.save(output_file)

import os
import re
import statistics
import openpyxl
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import numpy as np

# [Previous code remains unchanged up to the create_plots function]

def create_plots(stats, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    intervals = INTERVALS
    metrics = ['mse', 'iou']
    tokens = ['baseline', 'learned']
    colors = {'baseline': 'blue', 'learned': 'red'}

    # Get all unique chair IDs
    chair_ids = set()
    for interval in intervals:
        for view in VIEWS:
            for metric in metrics:
                chair_ids.update(stats[interval][view][metric].keys())

    # Create plots for each chair
    for chair_id in chair_ids:
        fig, axs = plt.subplots(2, 1, figsize=(10, 12), sharex=True)
        fig.suptitle(f'Metrics for Chair {chair_id}', fontsize=16)
        
        for idx, metric in enumerate(metrics):
            for token in tokens:
                means = []
                stds = []
                for interval in intervals:
                    interval_values = []
                    for view in VIEWS:
                        if chair_id in stats[interval][view][metric]:
                            value = stats[interval][view][metric][chair_id][token]
                            if '±' in value:
                                mean, _ = map(float, value.split('±'))
                                interval_values.append(mean)
                    if interval_values:
                        means.append(np.mean(interval_values))
                        stds.append(np.std(interval_values))
                    else:
                        means.append(np.nan)
                        stds.append(np.nan)
                
                means = np.array(means)
                stds = np.array(stds)
                
                axs[idx].plot(intervals, means, color=colors[token], label=token)
                axs[idx].fill_between(intervals, means-stds, means+stds, alpha=0.2, color=colors[token])
            
            axs[idx].set_ylabel(metric.upper())
            axs[idx].legend()
            axs[idx].grid(True, linestyle='--', alpha=0.7)

        axs[-1].set_xlabel('Interval')
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, f'{chair_id}_plot.png'))
        plt.close()

    # Create overall plot
    fig, axs = plt.subplots(2, 1, figsize=(10, 12), sharex=True)
    fig.suptitle('Overall Metrics', fontsize=16)
    
    for idx, metric in enumerate(metrics):
        for token in tokens:
            means = []
            stds = []
            for interval in intervals:
                interval_values = []
                for view in VIEWS:
                    for chair_id in stats[interval][view][metric]:
                        value = stats[interval][view][metric][chair_id][token]
                        if '±' in value:
                            mean, _ = map(float, value.split('±'))
                            interval_values.append(mean)
                if interval_values:
                    means.append(np.mean(interval_values))
                    stds.append(np.std(interval_values))
                else:
                    means.append(np.nan)
                    stds.append(np.nan)
            
            means = np.array(means)
            stds = np.array(stds)
            
            axs[idx].plot(intervals, means, color=colors[token], label=token)
            axs[idx].fill_between(intervals, means-stds, means+stds, alpha=0.2, color=colors[token])
        
        axs[idx].set_ylabel(metric.upper())
        axs[idx].legend()
        axs[idx].grid(True, linestyle='--', alpha=0.7)

    axs[-1].set_xlabel('Interval')
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'overall_plot.png'))
    plt.close()

# Main execution
base_dir = 'data/sd_1_5_output'
folder = base_dir.split('/')[-1]
output_dir = f'tables_plots/{folder}' #_binarized'
os.makedirs(output_dir, exist_ok=True)
excel_output = f'{output_dir}/{folder}_results.xlsx'
plot_output = f'{output_dir}/plots'

data = process_data(base_dir)
stats = calculate_stats(data)
create_excel(stats, excel_output)
create_plots(stats, plot_output)